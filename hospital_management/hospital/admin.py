from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template

from . models import Department,Doctor,Booking,Slider,About
# Register your models here.

admin.site.register(Department)
admin.site.register(Doctor)


def render_to_pdf(template_path, context_dict):
    template = get_template(template_path)
    html = template.render(context_dict)
    result = BytesIO()

    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)

    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None

def export_selected_to_pdf(modeladmin, request, queryset):
    template_path = 'pdf_template.html'
    context = {'bookings': queryset}
    pdf_data = render_to_pdf(template_path, context)

    if pdf_data:
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bookings.pdf"'
        return response

export_selected_to_pdf.short_description = "Export selected to PDF"



def export_selected_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'

    wb = Workbook()
    ws = wb.active

    # Write headers
    headers = ['Patient Name', 'Patient Phone', 'Symptoms', 'Doctor Name', 'Booking Date', 'Booked On']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws.column_dimensions[col_letter].width = 15

    # Write data
    for row_num, booking in enumerate(queryset, 2):
        ws.cell(row=row_num, column=1, value=booking.p_name)
        ws.cell(row=row_num, column=2, value=booking.p_phone)
        ws.cell(row=row_num, column=3, value=booking.symptoms)
        
        # Use str(booking.doc_name) to get the formatted representation
        ws.cell(row=row_num, column=4, value=str(booking.doc_name) if booking.doc_name else '')
        
        # Convert DateField to a formatted string
        ws.cell(row=row_num, column=5, value=booking.booking_date.strftime('%d-%m-%y') if booking.booking_date else '')
        
        # Convert DateField to a formatted string
        ws.cell(row=row_num, column=6, value=booking.booked_on.strftime('%d-%m-%y') if booking.booked_on else '')

    wb.save(response)
    return response

export_selected_to_excel.short_description = "Export selected to Excel"




class BookingAdmin(admin.ModelAdmin):
    list_display=('id','p_name','p_phone','symptoms','doc_name','booking_date','booked_on','status','token')
    list_filter = ('doc_name__doc_name', 'booking_date', 'status')
    actions = [export_selected_to_pdf, export_selected_to_excel]
    
admin.site.register(Booking,BookingAdmin)
admin.site.register(Slider)
admin.site.register(About)


